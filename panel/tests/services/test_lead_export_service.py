"""Unit tests for lead_export_service.

Pure tests — no DB. Build a workbook from synthetic lead dicts and verify
the openpyxl-parsed result has the expected headers, rows, and content.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.lead_export_service import (
    EXPORT_COLUMNS,
    build_leads_xlsx,
    export_filename,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load(buf: bytes):
    wb = load_workbook(BytesIO(buf), read_only=False)
    return wb.active


def _ic_direct_lead(**over) -> dict:
    base = {
        "id": 1,
        "name": "Ana Pérez",
        "phone": "+595981999001",
        "source": "infocasas",
        "is_direct_ic": True,
        "status": "interested",
        "created_at": datetime(2026, 5, 1, 12, 0, tzinfo=timezone.utc),
        "last_activity_at": datetime(2026, 5, 10, 16, 30, tzinfo=timezone.utc),
        "property_title": None,
        "property_city": None,
        "property_neighborhood": None,
        "property_price": None,
        "property_url": None,
        "ic_title": "Casa 3 dorm Asunción",
        "ic_city": "Asunción",
        "ic_price_sale": 150000,
        "ic_currency_sale": "USD",
        "ic_price_rent": None,
        "ic_currency_rent": None,
        "ic_url": "https://www.infocasas.com.py/propiedad/12345",
    }
    base.update(over)
    return base


def _whatsapp_lead(**over) -> dict:
    base = {
        "id": 2,
        "name": "Juan",
        "phone": "+595981999002",
        "source": "whatsapp",
        "is_direct_ic": None,
        "status": "new",
        "created_at": datetime(2026, 5, 2, 10, 0, tzinfo=timezone.utc),
        "last_activity_at": datetime(2026, 5, 2, 11, 0, tzinfo=timezone.utc),
        "property_title": "Departamento Villa Morra",
        "property_city": "Asunción",
        "property_neighborhood": "Villa Morra",
        "property_price": 95000,
        "property_operation": "sale",
        "property_url": "https://onnix.com.py/properties/777",
        "ic_title": None,
        "ic_city": None,
        "ic_price_sale": None,
        "ic_price_rent": None,
        "ic_currency_sale": None,
        "ic_currency_rent": None,
        "ic_url": None,
    }
    base.update(over)
    return base


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestHeaders:
    def test_first_row_matches_export_columns(self):
        ws = _load(build_leads_xlsx([]))
        actual = [cell.value for cell in ws[1]]
        assert actual == EXPORT_COLUMNS

    def test_columns_count_is_nine(self):
        """Nueve desde el 2026-08-24: entró «Consulta del cliente». El número
        solo, sin el test de abajo que la nombra, no vería un cambio de nombre
        ni un reemplazo — por eso van los dos."""
        assert len(EXPORT_COLUMNS) == 9

    def test_url_column_present(self):
        assert "URL propiedad" in EXPORT_COLUMNS

    def test_consulta_del_cliente_column_present(self):
        assert "Consulta del cliente" in EXPORT_COLUMNS


class TestSourceLabel:
    def test_infocasas_direct_renders_as_infocasas_directo(self):
        ws = _load(build_leads_xlsx([_ic_direct_lead()]))
        assert ws.cell(row=2, column=3).value == "infocasas_directo"

    def test_infocasas_indirect_renders_as_infocasas_reenviado(self):
        ws = _load(build_leads_xlsx([_ic_direct_lead(is_direct_ic=False)]))
        assert ws.cell(row=2, column=3).value == "infocasas_reenviado"

    def test_whatsapp_renders_verbatim(self):
        ws = _load(build_leads_xlsx([_whatsapp_lead()]))
        assert ws.cell(row=2, column=3).value == "whatsapp"

    def test_telegram_renders_verbatim(self):
        ws = _load(build_leads_xlsx([_whatsapp_lead(source="telegram")]))
        assert ws.cell(row=2, column=3).value == "telegram"


class TestPropertySummary:
    def test_uses_properties_table_when_present(self):
        ws = _load(build_leads_xlsx([_whatsapp_lead()]))
        cell = ws.cell(row=2, column=4).value
        assert "Departamento Villa Morra" in cell
        # El export lo abre el asesor y lo manda: el separador es el punto.
        assert "USD 95.000" in cell
        assert "Villa Morra" in cell

    def test_uses_ic_fields_when_no_property(self):
        ws = _load(build_leads_xlsx([_ic_direct_lead()]))
        cell = ws.cell(row=2, column=4).value
        assert "Casa 3 dorm Asunción" in cell
        assert "USD 150.000" in cell
        assert "Asunción" in cell

    def test_ic_rent_uses_currency_label(self):
        lead = _ic_direct_lead(
            ic_price_sale=None,
            ic_price_rent=2500000,
            ic_currency_rent="PYG",
        )
        ws = _load(build_leads_xlsx([lead]))
        cell = ws.cell(row=2, column=4).value
        assert "Gs 2.500.000" in cell
        assert "Alquiler" in cell

    def test_empty_property_renders_empty(self):
        lead = _whatsapp_lead(
            property_title=None,
            property_city=None,
            property_neighborhood=None,
            property_price=None,
        )
        ws = _load(build_leads_xlsx([lead]))
        # openpyxl roundtrips empty strings as None
        assert ws.cell(row=2, column=4).value in (None, "")


class TestPropertyURL:
    def test_uses_property_url_when_present(self):
        ws = _load(build_leads_xlsx([_whatsapp_lead()]))
        cell = ws.cell(row=2, column=5)
        assert cell.value == "https://onnix.com.py/properties/777"
        # When a URL is present it must also be a clickable hyperlink
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == "https://onnix.com.py/properties/777"

    def test_falls_back_to_ic_url(self):
        ws = _load(build_leads_xlsx([_ic_direct_lead()]))
        cell = ws.cell(row=2, column=5)
        assert cell.value == "https://www.infocasas.com.py/propiedad/12345"
        assert cell.hyperlink is not None

    def test_property_url_takes_priority_over_ic_url(self):
        lead = _whatsapp_lead(
            property_url="https://onnix.com.py/properties/111",
            ic_url="https://www.infocasas.com.py/propiedad/999",
        )
        ws = _load(build_leads_xlsx([lead]))
        assert ws.cell(row=2, column=5).value == "https://onnix.com.py/properties/111"

    def test_no_url_renders_empty_without_hyperlink(self):
        lead = _whatsapp_lead(property_url=None, ic_url=None)
        ws = _load(build_leads_xlsx([lead]))
        cell = ws.cell(row=2, column=5)
        assert cell.value in (None, "")
        assert cell.hyperlink is None


class TestRowMapping:
    def test_all_basic_columns_populated(self):
        lead = _ic_direct_lead()
        ws = _load(build_leads_xlsx([lead]))
        # New column order: name, phone, source, property summary, URL, status, created, last
        assert ws.cell(row=2, column=1).value == "Ana Pérez"
        assert ws.cell(row=2, column=2).value == "+595981999001"
        assert ws.cell(row=2, column=6).value == "interested"
        # Dates rendered in Paraguay time (UTC-3) — 12:00 UTC -> 09:00 / 16:30 UTC -> 13:30 (DST varies)
        # Just check the date portion is present.
        assert ws.cell(row=2, column=7).value.startswith("01/05/2026")
        assert ws.cell(row=2, column=8).value.startswith("10/05/2026")

    def test_multiple_leads_produce_multiple_rows(self):
        ws = _load(build_leads_xlsx([_ic_direct_lead(), _whatsapp_lead()]))
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_empty_leads_only_header(self):
        ws = _load(build_leads_xlsx([]))
        assert ws.max_row == 1


class TestExportFilename:
    def test_includes_tab_and_timestamp(self):
        name = export_filename("all", None, None)
        assert name.startswith("leads_all_")
        assert name.endswith(".xlsx")

    def test_includes_source_and_status_when_present(self):
        name = export_filename("all", "infocasas", "interested")
        assert "infocasas" in name
        assert "interested" in name

    def test_interested_tab_in_filename(self):
        name = export_filename("interested", None, None)
        assert "interested" in name
