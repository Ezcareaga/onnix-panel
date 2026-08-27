"""TDD — GET /leads/export

Verifies the route returns an xlsx file, respects tab/q/source/agent_id
filters when calling the service, and that the response includes the correct
content-type and disposition headers.

Vocabulary rewrite (2026-06-11, Chunk 2 LEADS-03 — plan
2026-06-11-leads-cola-trabajo-y-tono-bot): the legacy export vocabulary
(`tab=all|interested|agent_replied` + free-form `status` param) was removed.
`agent_replied` died as a filterable lead bucket with migration 018 (the
branch was dead code), and the page tabs speak M6.1
(`leads|interesados|asignados|sin_respuesta`). Export now flows through the
SAME WHERE-building as GET /leads (lead_service.list_leads_for_export →
lead_repo.get_by_tab), so the old asserts against `source=`/`status=` repo
kwargs no longer apply.

Mocks lead_service so tests do not need DB data.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from app.services.lead_export_service import EXPORT_COLUMNS

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SERVICE_TARGET = "app.routes.leads.lead_service.list_leads_for_export"


def _sample_lead(**over) -> dict:
    base = {
        "id": 1,
        "name": "Test Lead",
        "phone": "+595981999777",
        "source": "whatsapp",
        "is_direct_ic": None,
        "status": "new",
        "created_at": datetime(2026, 5, 1, 12, 0, tzinfo=timezone.utc),
        "last_activity_at": datetime(2026, 5, 2, 9, 0, tzinfo=timezone.utc),
        "property_title": "Casa centro",
        "property_city": "Asunción",
        "property_neighborhood": None,
        "property_price": 120000,
        "property_url": "https://onnix.com.py/properties/123",
        "ic_title": None,
        "ic_city": None,
        "ic_price_sale": None,
        "ic_price_rent": None,
        "ic_currency_sale": None,
        "ic_currency_rent": None,
        "ic_url": None,
    }
    base.update(over)
    return base


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

class TestExportAuth:
    async def test_unauthenticated_redirects_to_login(self, client):
        resp = await client.get("/leads/export")
        assert resp.status_code == 303
        assert "/login" in resp.headers["location"]

    async def test_user_role_gets_403(self, user_client):
        """Coherente con GET /leads (require_agent_or_admin): role 'user'
        no puede ver leads → tampoco exportarlos."""
        resp = await user_client.get("/leads/export")
        assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Response shape
# ---------------------------------------------------------------------------

class TestExportResponse:
    async def test_returns_xlsx_content_type(self, admin_client):
        with patch(SERVICE_TARGET, new=AsyncMock(return_value=[])):
            resp = await admin_client.get("/leads/export")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(XLSX_MIME)

    async def test_content_disposition_attachment_with_xlsx_filename(self, admin_client):
        with patch(SERVICE_TARGET, new=AsyncMock(return_value=[])):
            resp = await admin_client.get("/leads/export")
        disposition = resp.headers["content-disposition"]
        assert "attachment" in disposition
        assert ".xlsx" in disposition

    async def test_body_is_a_valid_xlsx_with_expected_headers(self, admin_client):
        with patch(SERVICE_TARGET, new=AsyncMock(return_value=[_sample_lead()])):
            resp = await admin_client.get("/leads/export")
        wb = load_workbook(BytesIO(resp.content), read_only=False)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        assert header_row == EXPORT_COLUMNS
        # 1 header + 1 data row from the mocked lead
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "Test Lead"

    async def test_url_column_renders_property_link(self, admin_client):
        with patch(SERVICE_TARGET, new=AsyncMock(return_value=[_sample_lead()])):
            resp = await admin_client.get("/leads/export")
        wb = load_workbook(BytesIO(resp.content), read_only=False)
        ws = wb.active
        url_cell = ws.cell(row=2, column=5)
        assert url_cell.value == "https://onnix.com.py/properties/123"


# ---------------------------------------------------------------------------
# Filter pass-through (M6.1 vocabulary — same WHERE-building as GET /leads)
# ---------------------------------------------------------------------------

class TestExportFilters:
    async def test_tab_and_filters_passed_to_service(self, admin_client):
        mock_service = AsyncMock(return_value=[])
        with patch(SERVICE_TARGET, new=mock_service):
            resp = await admin_client.get(
                "/leads/export?tab=interesados&q=Pena&source=infocasas&agent_id=7"
            )
        assert resp.status_code == 200
        _, kwargs = mock_service.call_args
        assert kwargs["tab"] == "interesados"
        assert kwargs["q"] == "Pena"
        assert kwargs["source"] == "infocasas"
        assert kwargs["agent_id"] == 7
        assert kwargs["agent_filter"] is None

    async def test_no_filter_defaults_to_tab_leads(self, admin_client):
        mock_service = AsyncMock(return_value=[])
        with patch(SERVICE_TARGET, new=mock_service):
            resp = await admin_client.get("/leads/export")
        assert resp.status_code == 200
        _, kwargs = mock_service.call_args
        assert kwargs["tab"] == "leads"
        assert kwargs["q"] is None
        assert kwargs["source"] is None
        assert kwargs["agent_id"] is None

    @pytest.mark.parametrize("legacy_tab", ["all", "interested", "agent_replied"])
    async def test_legacy_tab_values_normalize_to_leads(self, admin_client, legacy_tab):
        """Vocabulario viejo eliminado: tab desconocido → default 'leads'
        (mismo comportamiento que GET /leads). 'agent_replied' era branch
        muerto desde mig 018."""
        mock_service = AsyncMock(return_value=[])
        with patch(SERVICE_TARGET, new=mock_service):
            resp = await admin_client.get(f"/leads/export?tab={legacy_tab}")
        assert resp.status_code == 200
        _, kwargs = mock_service.call_args
        assert kwargs["tab"] == "leads"

    async def test_agent_role_forced_to_own_asignados(self, agent_client):
        """Mismo enforcement de rol que GET /leads: un agent exporta SOLO
        su bucket de asignados, ignorando tab/agent_id del query string."""
        mock_service = AsyncMock(return_value=[])
        with patch(SERVICE_TARGET, new=mock_service):
            resp = await agent_client.get("/leads/export?tab=leads&agent_id=999")
        assert resp.status_code == 200
        _, kwargs = mock_service.call_args
        assert kwargs["tab"] == "asignados"
        assert kwargs["agent_filter"] is not None
        assert kwargs["agent_id"] is None
