"""Chunks 1-2 — Leads como cola de trabajo (plan 2026-06-11-leads-cola-trabajo-y-tono-bot).

Covers the lead row improvements:
    LEADS-05 — source badges (source_badge.html partial)
    LEADS-02 — real interest column (IC > linked property > search_context > "—")
    LEADS-01 — waiting urgency (semaforo verde/ambar/rojo + oldest-first order)
    LEADS-03 — search (q unaccent name/phone) + filters (source, agent_id),
               preserved in tabs/pagination/export; export on M6.1 vocabulary

Row-content tests render partials/lead_item.html through the existing
POST /leads/{id}/status swap endpoint (same-status no-op) instead of paging
through GET /leads — the dev DB has hundreds of rows per tab, so asserting on
page 1 of the full table would be pagination-fragile.

Test contacts use the +5959817 prefix → cleaned up by conftest session cleanup.
"""
import random
from datetime import datetime, timedelta, timezone

import pytest
from sqlalchemy import text as sa_text

from app.models.contact import Contact
from app.models.conversation import Conversation


def _suffix() -> int:
    return random.randint(100_000, 999_000)


def _make_contact(db, phone: str, **kw):
    defaults = dict(
        name="Workqueue Probe",
        source="manual",
        status="new",
        agent_user_id=None,
        created_at=datetime.now(timezone.utc),
        last_activity_at=datetime.now(timezone.utc),
    )
    defaults.update(kw)
    c = Contact(phone=phone, **defaults)
    db.add(c)
    return c


async def _row_html(admin_client, contact) -> str:
    """Render partials/lead_item.html for one contact via the status swap
    endpoint (posting the current status is a no-op in lead_service)."""
    resp = await admin_client.post(
        f"/leads/{contact.id}/status", data={"status": contact.status},
    )
    assert resp.status_code == 200, resp.text
    return resp.text


# ---------------------------------------------------------------------------
# Task 1 — LEADS-05 source badges
# ---------------------------------------------------------------------------

class TestSourceBadge:
    async def test_lead_row_shows_source_badge(self, admin_client, db):
        base = _suffix()
        # Carril B3: las fuentes dejan de tener matiz propio. Eran cinco
        # colores compitiendo con el estado del lead, que es lo unico de la
        # fila que de verdad pide algo. Todas van en la variante `default` y
        # las siglas pasan a nombres, que es lo que se lee sin traducir.
        cases = [
            ("whatsapp", "WhatsApp"),
            ("telegram", "Telegram"),
            ("infocasas", "InfoCasas"),
            ("vista_publica", "Portal"),
            ("manual", "Manual"),
            ("import:crm", "Importado"),
        ]
        contacts = []
        for i, (source, _) in enumerate(cases):
            contacts.append(
                _make_contact(db, f"+5959817{base + i}", source=source)
            )
        await db.commit()
        for c in contacts:
            await db.refresh(c)

        for contact, (source, etiqueta) in zip(contacts, cases):
            html = await _row_html(admin_client, contact)
            assert f">{etiqueta}<" in html, (
                f"source={source!r}: falta la etiqueta {etiqueta!r} en la fila"
            )
            assert "badge badge--default" in html, (
                f"source={source!r}: la fuente no usa la variante `default`"
            )


# ---------------------------------------------------------------------------
# Task 2 — LEADS-02 real interest column
# ---------------------------------------------------------------------------

class TestInterestColumn:
    @pytest.fixture
    async def ic_lead(self, db):
        """IC contact + matching infocasas_properties row (manual teardown —
        conftest cleanup does not cover infocasas_properties)."""
        ref = "WQIC01"
        title = "Casa WQ Probe en Lambare"
        phone = f"+5959817{_suffix()}"
        await db.execute(
            sa_text("DELETE FROM infocasas_properties WHERE infocasas_ref = :ref"),
            {"ref": ref},
        )
        await db.execute(
            sa_text(
                "INSERT INTO infocasas_properties "
                "(infocasas_id, infocasas_ref, title, city, is_active) "
                "VALUES ('WQIC001', :ref, :title, 'Lambare', true)"
            ),
            {"ref": ref, "title": title},
        )
        contact = _make_contact(
            db, phone, source="infocasas", infocasas_ref=ref,
        )
        await db.commit()
        await db.refresh(contact)
        yield {"contact": contact, "title": title}
        await db.execute(
            sa_text("DELETE FROM infocasas_properties WHERE infocasas_ref = :ref"),
            {"ref": ref},
        )
        await db.commit()

    async def test_lead_row_interest_from_ic(self, admin_client, ic_lead):
        html = await _row_html(admin_client, ic_lead["contact"])
        assert ic_lead["title"][:50] in html, (
            "IC consulta title must win the interest column for IC leads"
        )
        assert "No especifico" not in html

    async def test_lead_row_interest_from_search_context(self, admin_client, db):
        """No IC, no linked property → summary of the LATEST conversation's
        search_context (created_at DESC) renders in the interest column."""
        contact = _make_contact(db, f"+5959817{_suffix()}", source="whatsapp")
        await db.commit()
        await db.refresh(contact)

        now = datetime.now(timezone.utc)
        older = Conversation(
            contact_id=contact.id,
            channel="whatsapp",
            status="active",
            created_at=now - timedelta(days=2),
            search_context={"filtros": {
                "tipo": "casa", "operacion": None,
                "ciudad": "Asuncion", "barrio": None,
            }},
        )
        newer = Conversation(
            contact_id=contact.id,
            channel="whatsapp",
            status="active",
            created_at=now - timedelta(hours=1),
            search_context={"filtros": {
                "tipo": "departamento", "operacion": "alquiler",
                "ciudad": "Luque", "barrio": None,
            }},
        )
        db.add_all([older, newer])
        await db.commit()

        html = await _row_html(admin_client, contact)
        assert "Departamento · Alquiler · Luque" in html, (
            "interest column must summarize the latest conversation's "
            "search_context (filtros nested keys)"
        )
        assert "Asuncion" not in html, "older conversation context must not win"
        assert "No especifico" not in html

    async def test_lead_row_interest_dash_when_nothing(self, admin_client, db):
        """No IC, no property, no search_context → '—' (never 'No especifico')."""
        contact = _make_contact(db, f"+5959817{_suffix()}", source="whatsapp")
        await db.commit()
        await db.refresh(contact)

        html = await _row_html(admin_client, contact)
        assert "No especifico" not in html, (
            "the misspelled 'No especifico' placeholder must be gone"
        )
        assert "—" in html


class TestSummarizeSearchContext:
    def test_summarize_search_context_variants(self):
        from app.services.lead_service import summarize_search_context

        # 1. Full real-shape context (nulls explicit, extra keys ignored)
        full = {"filtros": {
            "tipo": "departamento", "operacion": "alquiler",
            "ciudad": "Luque", "barrio": "Villa Morra",
            "precio_max": None, "dormitorios": None,
        }}
        assert summarize_search_context(full) == (
            "Departamento · Alquiler · Luque · Villa Morra"
        )

        # 2. Only tipo — every other key null
        solo_tipo = {"filtros": {
            "tipo": "casa", "operacion": None, "ciudad": None, "barrio": None,
        }}
        assert summarize_search_context(solo_tipo) == "Casa"

        # 3. Only ciudad
        solo_ciudad = {"filtros": {
            "tipo": None, "operacion": None,
            "ciudad": "San Lorenzo", "barrio": None,
        }}
        assert summarize_search_context(solo_ciudad) == "San Lorenzo"

        # 4. All-null filtros → None
        vacio = {"filtros": {
            "tipo": None, "operacion": None, "ciudad": None, "barrio": None,
        }}
        assert summarize_search_context(vacio) is None

        # 5. Degenerate inputs → None
        assert summarize_search_context(None) is None
        assert summarize_search_context({}) is None
        assert summarize_search_context({"filtros": None}) is None

        # 6. JSONB delivered as a JSON string must be deserialized
        as_string = (
            '{"filtros": {"tipo": "casa", "operacion": "venta",'
            ' "ciudad": null, "barrio": null}}'
        )
        assert summarize_search_context(as_string) == "Casa · Venta"
        assert summarize_search_context("not-json{") is None

        # 7. Capitalization preserves inner uppercase (no .capitalize() squash)
        mixed = {"filtros": {"tipo": None, "operacion": None,
                             "ciudad": "Villa Elisa", "barrio": None}}
        assert summarize_search_context(mixed) == "Villa Elisa"


# ---------------------------------------------------------------------------
# Task 3 — LEADS-01 waiting urgency (semaforo + oldest-first order)
# ---------------------------------------------------------------------------

class TestWaitingBuckets:
    def test_waiting_badge_buckets(self):
        from app.services.lead_service import compute_waiting, format_relative_es

        now = datetime(2026, 6, 11, 12, 0, 0, tzinfo=timezone.utc)

        def at(**delta):
            return now - timedelta(**delta)

        # verde: < 1h waiting, status requiring human action
        bucket, since = compute_waiting("new", None, at(minutes=25), now=now)
        assert bucket == "verde" and since == at(minutes=25)

        # ambar: 1-24h
        bucket, _ = compute_waiting("bot_replied", None, at(hours=3), now=now)
        assert bucket == "ambar"

        # rojo: > 24h
        bucket, _ = compute_waiting("no_response", None, at(days=2), now=now)
        assert bucket == "rojo"

        # interested SIN agente asignado → needs action
        bucket, _ = compute_waiting("interested", None, at(hours=2), now=now)
        assert bucket == "ambar"

        # interested CON agente → no urgency badge
        assert compute_waiting("interested", 7, at(days=3), now=now) == (None, None)

        # statuses fuera del cohorte → no badge
        assert compute_waiting("closed", None, at(days=9), now=now) == (None, None)
        assert compute_waiting("agent_replied", None, at(days=9), now=now) == (None, None)

        # sin last_activity_at → no badge
        assert compute_waiting("new", None, None, now=now) == (None, None)

        # relative-time labels in Spanish
        assert format_relative_es(at(minutes=25), now=now) == "hace 25 min"
        assert format_relative_es(at(hours=3), now=now) == "hace 3 h"
        assert format_relative_es(at(days=2), now=now) == "hace 2 días"
        assert format_relative_es(at(days=1), now=now) == "hace 1 día"
        assert format_relative_es(at(seconds=10), now=now) == "hace 1 min"

    async def test_waiting_said_in_words_in_row(self, admin_client, db):
        """Decision 4 — la espera se dice ("Esperando hace 3 dias"), no se pinta.

        El punto de color se fue con la temperatura: mientras convivieron, el
        mismo rojo significaba "lead caliente" (bueno) y "espera hace mucho"
        (malo). El bucket sigue existiendo y sigue ordenando la cola.
        """
        contact = _make_contact(
            db, f"+5959817{_suffix()}",
            source="whatsapp", status="no_response",
            last_activity_at=datetime.now(timezone.utc) - timedelta(days=3),
        )
        await db.commit()
        await db.refresh(contact)

        html = await _row_html(admin_client, contact)
        assert "Esperando hace 3 días" in html, (
            "la fila debe decir la espera en palabras"
        )
        # Ni punto ni color: los tres tokens del semaforo viejo, buscados en el
        # markup con las llaves de clase que los rodeaban.
        for painted in ('rounded-full bg-green-500', 'rounded-full bg-amber-400',
                        'rounded-full bg-red-500', 'text-red-500 font-semibold'):
            assert painted not in html, f"el semaforo de color volvio: {painted}"


class TestLastContactRelative:
    """Limpieza (c) — 'Último contacto' unificado a relativo en la tabla."""

    def test_enrich_adds_last_contact_fields(self):
        from app.services.lead_service import _enrich_lead_row

        now = datetime(2026, 6, 12, 12, 0, 0, tzinfo=timezone.utc)
        ts = now - timedelta(hours=2)
        row = {
            "status": "agent_replied", "agent_user_id": 7,
            "last_activity_at": ts, "created_at": ts,
            "source": "whatsapp", "consulta_date": None,
            "search_context": None,
        }
        _enrich_lead_row(row, now=now)
        assert row["last_contact_at"] == ts
        assert row["last_contact_label"] == "hace 2 h"

    def test_enrich_ic_uses_consulta_date(self):
        from app.services.lead_service import _enrich_lead_row

        now = datetime(2026, 6, 12, 12, 0, 0, tzinfo=timezone.utc)
        consulta = now - timedelta(days=1)
        row = {
            "status": "new", "agent_user_id": None,
            "last_activity_at": now - timedelta(hours=1),
            "created_at": now - timedelta(days=5),
            "source": "infocasas", "consulta_date": consulta,
            "search_context": None,
        }
        _enrich_lead_row(row, now=now)
        assert row["last_contact_at"] == consulta, (
            "IC leads must keep showing the consulta date"
        )
        assert row["last_contact_label"] == "hace 1 día"

    def test_enrich_falls_back_to_created_at(self):
        from app.services.lead_service import _enrich_lead_row

        now = datetime(2026, 6, 12, 12, 0, 0, tzinfo=timezone.utc)
        created = now - timedelta(days=3)
        row = {
            "status": "new", "agent_user_id": None,
            "last_activity_at": None, "created_at": created,
            "source": "manual", "consulta_date": None,
            "search_context": None,
        }
        _enrich_lead_row(row, now=now)
        assert row["last_contact_at"] == created
        assert row["last_contact_label"] == "hace 3 días"

    async def test_row_renders_relative_last_contact_with_title(
        self, admin_client, db,
    ):
        """La celda 'Último contacto' muestra 'hace X' con title=fecha completa.

        status=agent_replied (sin waiting badge) para que el 'hace 2 h'
        asertado solo pueda venir de la celda de último contacto.
        """
        contact = _make_contact(
            db, f"+5959817{_suffix()}", source="whatsapp",
            status="agent_replied",
            last_activity_at=datetime.now(timezone.utc) - timedelta(hours=2),
        )
        await db.commit()
        await db.refresh(contact)

        html = await _row_html(admin_client, contact)
        assert "hace 2 h" in html, "last contact cell must render relative time"
        # title= con la fecha completa (dd/mm/yyyy via filtro pyt)
        year = str(datetime.now(timezone.utc).year)
        assert f'title="' in html and year in html, (
            "relative time must carry a title= with the full date"
        )


class TestSinAsignarChip:
    """'Sin asignar' exige accion, y eso se dice con peso, no con ambar.

    Era un chip ambar — un tercer matiz saturado en la misma fila donde ya
    competian el estado y la fuente. Desde el carril B3 usa la variante
    `strong`, que es lo mas oscuro de la pantalla: el mismo mensaje sin
    gastar un color."""

    async def test_row_unassigned_shows_strong_badge(self, admin_client, db):
        contact = _make_contact(db, f"+5959817{_suffix()}", source="whatsapp")
        await db.commit()
        await db.refresh(contact)

        html = await _row_html(admin_client, contact)
        assert "Sin asignar" in html
        chip = html.split("Sin asignar", 1)[0].rsplit("<span", 1)[1]
        assert "badge--strong" in chip, "'Sin asignar' tiene que exigir accion"
        assert "bg-amber-" not in chip, "el ambar volvio: es un tercer matiz saturado"
        assert "italic" not in chip, "el placeholder gris en italica tiene que estar muerto"


class TestSinRespuestaOrdering:
    async def test_tab_sin_respuesta_orders_oldest_first(self, db):
        """LEADS-01 — 'sin_respuesta' ordena por last_activity_at ASC NULLS
        LAST (más tiempo esperando primero). Es un backlog de recuperación:
        acá el ASC se queda. La bandeja «Nuevos» va al revés — ver
        TestNuevosOrdering más abajo.

        Service-level (per_page grande) — la página 1 del route es
        pagination-fragile con cientos de filas reales en dev.
        """
        from app.services.lead_service import lead_service

        base = _suffix()
        now = datetime.now(timezone.utc)
        spec = [
            ("old", now - timedelta(days=40)),
            ("mid", now - timedelta(days=39)),
            ("new", now - timedelta(days=38)),
        ]
        ids = {}
        for i, (label, ts) in enumerate(spec):
            c = _make_contact(
                db, f"+5959817{base + i}",
                source="whatsapp", status="no_response",
                created_at=ts, last_activity_at=ts,
            )
            await db.commit()
            await db.refresh(c)
            ids[label] = c.id

        rows, _total = await lead_service.list_leads_by_tab(
            db, "sin_respuesta", page=1, per_page=5000,
        )
        order = [r["id"] for r in rows]
        for label in ids:
            assert ids[label] in order, f"{label} contact missing from tab"
        assert order.index(ids["old"]) < order.index(ids["mid"]) < order.index(ids["new"]), (
            "sin_respuesta must order oldest last_activity_at first"
        )


class TestNuevosOrdering:
    """«Nuevos» es bandeja de entrada: lo más reciente primero (decisión de Ez,
    2026-08-24). Invierte el ASC que había puesto 18aad2c.

    Medido en producción antes del cambio: de 9 leads en la pestaña, el más
    reciente salía octavo. Estos tests fijan la dirección del orden y —el
    segundo— el efecto que más dolía: que una consulta nueva sobre un lead
    viejo lo hundiera al fondo en vez de subirlo.
    """

    async def test_tab_leads_orders_newest_first(self, db):
        from app.services.lead_service import lead_service

        base = _suffix()
        now = datetime.now(timezone.utc)
        older = _make_contact(
            db, f"+5959817{base}", status="new",
            created_at=now - timedelta(days=40),
            last_activity_at=now - timedelta(days=40),
        )
        newer = _make_contact(
            db, f"+5959817{base + 1}", status="bot_replied",
            created_at=now - timedelta(days=39),
            last_activity_at=now - timedelta(days=39),
        )
        await db.commit()
        await db.refresh(older)
        await db.refresh(newer)

        rows, _total = await lead_service.list_leads_by_tab(
            db, "leads", page=1, per_page=5000,
        )
        order = [r["id"] for r in rows]
        assert older.id in order and newer.id in order
        assert order.index(newer.id) < order.index(older.id), (
            "«Nuevos» tiene que mostrar el lead más reciente primero"
        )

    async def test_nueva_consulta_sube_el_lead_en_vez_de_hundirlo(self, db):
        """El lead viejo que vuelve a consultar (last_activity_at se actualiza)
        tiene que subir al tope, no caer al fondo. Con el ASC anterior pasaba
        exactamente lo contrario: por eso el 23194, del 20/08, salía noveno."""
        from app.services.lead_service import lead_service

        base = _suffix()
        now = datetime.now(timezone.utc)
        viejo_que_reconsulta = _make_contact(
            db, f"+5959817{base}", status="new",
            created_at=now - timedelta(days=40),
            last_activity_at=now - timedelta(minutes=1),  # volvió a escribir
        )
        reciente_sin_reconsulta = _make_contact(
            db, f"+5959817{base + 1}", status="new",
            created_at=now - timedelta(hours=2),
            last_activity_at=now - timedelta(hours=2),
        )
        await db.commit()
        await db.refresh(viejo_que_reconsulta)
        await db.refresh(reciente_sin_reconsulta)

        rows, _total = await lead_service.list_leads_by_tab(
            db, "leads", page=1, per_page=5000,
        )
        order = [r["id"] for r in rows]
        assert order.index(viejo_que_reconsulta.id) < order.index(
            reciente_sin_reconsulta.id
        ), "una consulta nueva tiene que subir el lead, no hundirlo"

    async def test_sin_respuesta_conserva_el_asc(self, db):
        """La otra rama NO se movió: sin_respuesta dejó de compartir el `in
        (...)` con leads y tiene que seguir siendo el más viejo primero."""
        from app.services.lead_service import lead_service

        base = _suffix()
        now = datetime.now(timezone.utc)
        viejo = _make_contact(
            db, f"+5959817{base}", source="whatsapp", status="no_response",
            created_at=now - timedelta(days=50),
            last_activity_at=now - timedelta(days=50),
        )
        nuevo = _make_contact(
            db, f"+5959817{base + 1}", source="whatsapp", status="no_response",
            created_at=now - timedelta(days=49),
            last_activity_at=now - timedelta(days=49),
        )
        await db.commit()
        await db.refresh(viejo)
        await db.refresh(nuevo)

        rows, _total = await lead_service.list_leads_by_tab(
            db, "sin_respuesta", page=1, per_page=5000,
        )
        order = [r["id"] for r in rows]
        assert order.index(viejo.id) < order.index(nuevo.id), (
            "sin_respuesta sigue siendo cola de trabajo: el más viejo primero"
        )


# ---------------------------------------------------------------------------
# Chunk 2 — LEADS-03 búsqueda + filtros + export coherente
# ---------------------------------------------------------------------------

class TestLeadsSearch:
    async def test_leads_search_by_name_unaccent(self, admin_client, db):
        """q='... Pena' must find 'Peña' (unaccent both sides — regla 7)."""
        marker = f"Wqfiltro{_suffix()}"
        target = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} Peña", source="whatsapp",
        )
        decoy = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} Pino", source="whatsapp",
        )
        await db.commit()
        await db.refresh(target)
        await db.refresh(decoy)

        resp = await admin_client.get(
            "/leads", params={"tab": "leads", "q": f"{marker} Pena"},
        )
        assert resp.status_code == 200
        assert target.phone in resp.text, (
            "search 'Pena' must match 'Peña' via unaccent"
        )
        assert decoy.phone not in resp.text

    async def test_leads_search_by_phone(self, admin_client, db):
        """q with a phone substring matches via LIKE on c.phone."""
        phone = f"+5959818{_suffix()}"
        contact = _make_contact(db, phone, name="Wq Phone Probe", source="manual")
        await db.commit()
        await db.refresh(contact)

        resp = await admin_client.get(
            "/leads", params={"tab": "leads", "q": phone.lstrip("+")},
        )
        assert resp.status_code == 200
        assert phone in resp.text


class TestLeadsFilters:
    async def test_leads_filter_by_source(self, admin_client, db):
        marker = f"Wqsrc{_suffix()}"
        tg = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} TG", source="telegram",
        )
        wa = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} WA", source="whatsapp",
        )
        await db.commit()
        await db.refresh(tg)
        await db.refresh(wa)

        resp = await admin_client.get(
            "/leads", params={"tab": "leads", "q": marker, "source": "telegram"},
        )
        assert resp.status_code == 200
        assert tg.phone in resp.text
        assert wa.phone not in resp.text, "source filter must exclude other sources"

    async def test_leads_filter_by_agent(self, admin_client, db):
        from app.models.user import User

        marker = f"Wqagt{_suffix()}"
        n = _suffix()
        u1 = User(
            email=f"pytest_wqagent{n}a@onnixtest.com", name="WQ Agent A",
            role="agent", password_hash="x", is_active=True,
        )
        u2 = User(
            email=f"pytest_wqagent{n}b@onnixtest.com", name="WQ Agent B",
            role="agent", password_hash="x", is_active=True,
        )
        db.add_all([u1, u2])
        await db.flush()
        c1 = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} A",
            agent_user_id=u1.id, status="interested",
        )
        c2 = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} B",
            agent_user_id=u2.id, status="interested",
        )
        await db.commit()
        await db.refresh(c1)
        await db.refresh(c2)

        resp = await admin_client.get(
            "/leads",
            params={"tab": "asignados", "q": marker, "agent_id": u1.id},
        )
        assert resp.status_code == 200
        assert c1.phone in resp.text
        assert c2.phone not in resp.text, "agent_id filter must exclude other agents"

    async def test_leads_invalid_agent_id_ignored(self, admin_client):
        """Garbage agent_id must not 422/500 — sanitized to None."""
        resp = await admin_client.get(
            "/leads", params={"tab": "asignados", "agent_id": "abc"},
        )
        assert resp.status_code == 200


class TestLeadsFilterPreservation:
    async def test_leads_filters_preserved_in_pagination_links(self, admin_client, db):
        """Pagination + tab links + export button carry q/source/agent_id."""
        marker = f"Wqpag{_suffix()}"
        for i in range(26):  # per_page=25 → forces page 2 within the filter
            _make_contact(
                db, f"+5959818{_suffix()}", name=f"{marker} {i:02d}",
                source="whatsapp",
            )
        await db.commit()

        resp = await admin_client.get(
            "/leads", params={"tab": "leads", "q": marker, "source": "whatsapp"},
        )
        assert resp.status_code == 200
        html = resp.text

        # Pagination → next page keeps the filters
        assert f"?tab=leads&page=2&q={marker}&source=whatsapp" in html, (
            "pagination links must preserve q/source"
        )
        # Tabs keep the filters
        assert f"/leads?tab=interesados&q={marker}&source=whatsapp" in html, (
            "tab links must preserve q/source"
        )
        assert f"/leads?tab=sin_respuesta&q={marker}&source=whatsapp" in html
        # Export button keeps tab + filters (M6.1 vocabulary)
        assert f"/leads/export?tab=leads&q={marker}&source=whatsapp" in html, (
            "export link must carry the active tab + filters"
        )
        # Limpiar link visible with active filters
        assert "Limpiar" in html


# ---------------------------------------------------------------------------
# Chunk 3 — LEADS-04 una sola fila, que abajo de 768px es card (carril G)
# ---------------------------------------------------------------------------

class TestMobileCards:
    async def test_leads_single_item_per_lead(self, admin_client, db):
        """GET /leads renderiza UN nodo por lead, no dos.

        Habia dos plantillas sobre los mismos datos —la fila y una card
        mobile— y la card no podia cambiar el estado, ni asignar asesor, ni
        abrir la conversacion. Ahora es una sola y la reacomoda el CSS, asi
        que las capacidades ya no pueden divergir: si estan en el markup,
        estan en las dos formas.
        """
        marker = f"Wqcard{_suffix()}"
        phone = f"+5959818{_suffix()}"
        contact = _make_contact(
            db, phone, name=f"{marker} Probe", source="whatsapp", status="new",
            last_activity_at=datetime.now(timezone.utc) - timedelta(hours=3),
        )
        await db.commit()
        await db.refresh(contact)

        resp = await admin_client.get(
            "/leads", params={"tab": "leads", "q": marker},
        )
        assert resp.status_code == 200
        html = resp.text

        # Un solo nodo por lead: la card mobile paralela no existe mas.
        assert html.count(f'id="lead-row-{contact.id}"') == 1, (
            "el lead debe renderizarse una sola vez"
        )
        assert 'id="lead-card-' not in html, (
            "volvio la card mobile paralela — con ella vuelve la divergencia"
        )
        assert "leads-table" in html, (
            "la tabla necesita la clase que la reacomoda como card <768px"
        )

        card_html = html.split(f'id="lead-row-{contact.id}"', 1)[1].split("</tr>", 1)[0]
        assert f"{marker} Probe" in card_html, "card must show the lead name"
        assert f"/contacts/{contact.id}" in card_html, (
            "card name must link to the contact detail"
        )
        assert ">WhatsApp<" in card_html, "card must render the source badge"
        assert ">Nuevo<" in card_html, "card must render the status badge"
        assert "Esperando hace 3 h" in card_html, (
            "card must say the waiting time in words"
        )
        assert "—" in card_html, (
            "card must render the interest line ('—' when nothing known)"
        )
        wa_phone = phone.lstrip("+")
        assert f"wa.me/{wa_phone}" in card_html, (
            "card must include the WhatsApp deep-link when phone exists"
        )
        # Admin sees assignment state — unassigned lead shows the chip.
        assert "Sin asignar" in card_html, (
            "admin card must show 'Sin asignar' chip when no agent assigned"
        )

        # Las capacidades que la card vieja NO tenia, y que ahora vienen en el
        # mismo markup que la fila: cambiar estado, asignar asesor y abrir la
        # conversacion. Sin ellas la fusion no sirvio de nada.
        assert f'hx-post="/leads/{contact.id}/status"' in card_html, (
            "el lead debe poder cambiar de estado desde el celular"
        )
        assert f'hx-post="/leads/{contact.id}/agent-assign"' in card_html, (
            "el admin debe poder asignar asesor desde el celular"
        )
        assert "Sin conversación" in card_html or "/conversations/" in card_html, (
            "la conversacion del lead debe ser alcanzable desde el celular"
        )

    async def test_leads_tabs_bar_scrolls_horizontally_on_mobile(self, admin_client):
        """QA staging: a 390px la pagina scrolleaba horizontal (scrollWidth 571)
        porque la barra de tabs no scrollea internamente. La barra debe tener
        overflow-x-auto y los tabs deben comportarse como cinta (shrink-0 +
        whitespace-nowrap) dentro del contenedor."""
        resp = await admin_client.get("/leads", params={"tab": "leads"})
        assert resp.status_code == 200
        html = resp.text

        # The tabs bar is the last <div> opened before the first tab link.
        head = html.split('href="/leads?tab=leads', 1)[0]
        tabs_div = head.rsplit("<div", 1)[1].split(">", 1)[0]
        assert "border-b" in tabs_div, "did not locate the tabs bar div"
        assert "overflow-x-auto" in tabs_div, (
            "tabs bar must scroll horizontally (overflow-x-auto) on mobile"
        )

        # Each tab must stay on one line and not shrink, so the bar scrolls
        # as a ribbon instead of overflowing the page.
        tabs_block = html.split('href="/leads?tab=leads', 1)[1].split("</div>", 1)[0]
        for tab in ("leads", "interesados", "sin_respuesta", "asignados"):
            anchor = html.split(f'href="/leads?tab={tab}', 1)[1].split("</a>", 1)[0]
            assert "shrink-0" in anchor, f"tab {tab} must have shrink-0"
            assert "whitespace-nowrap" in anchor, (
                f"tab {tab} must have whitespace-nowrap"
            )
        assert tabs_block  # sanity

    async def test_leads_htmx_partial_carries_the_rows(self, admin_client, db):
        """El partial de HX-Request (el swap del SSE) trae las filas.

        Antes tenia que traer las dos vistas o el refresh las desincronizaba;
        con una sola plantilla el problema no existe.
        """
        marker = f"Wqcard{_suffix()}"
        contact = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} Swap", source="whatsapp",
        )
        await db.commit()
        await db.refresh(contact)

        resp = await admin_client.get(
            "/leads", params={"tab": "leads", "q": marker},
            headers={"HX-Request": "true"},
        )
        assert resp.status_code == 200
        assert f'id="lead-row-{contact.id}"' in resp.text, (
            "HTMX partial must include the table rows"
        )


class TestQuickFilterChips:
    """La fila de chips se borro en el carril D.

    Los tres chips reproducian combinaciones de tab+filtro que estaban un
    renglon mas arriba, y el tercero mentia: se llamaba "sin asignar" y
    linkeaba al tab entero. Lo que el test cuida es que borrarlos no se
    haya llevado ningun destino.
    """

    async def test_los_chips_no_repiten_la_fila_de_tabs(self, admin_client):
        resp = await admin_client.get("/leads")
        assert resp.status_code == 200
        html = resp.text
        assert "Filtros r\u00e1pidos" not in html
        for destino in (
            "/leads?tab=sin_respuesta",
            "/leads?tab=leads",
            "/leads?tab=interesados",
        ):
            assert f'href="{destino}' in html, (
                f"{destino} dejo de ser alcanzable desde la fila de tabs"
            )


class TestLeadsExportFilters:
    async def test_leads_export_respects_filters(self, admin_client, db):
        """Export with M6.1 tab + q returns an xlsx with only matching rows."""
        from io import BytesIO

        from openpyxl import load_workbook

        marker = f"Wqexp{_suffix()}"
        target = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} Target",
            source="whatsapp", status="interested",
        )
        decoy = _make_contact(
            db, f"+5959818{_suffix()}", name="Wqexpdecoy Otro",
            source="whatsapp", status="interested",
        )
        await db.commit()
        await db.refresh(target)
        await db.refresh(decoy)

        resp = await admin_client.get(
            "/leads/export", params={"tab": "interesados", "q": marker},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "interesados" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.content), read_only=False)
        ws = wb.active
        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert f"{marker} Target" in names, "filtered export must include the match"
        assert "Wqexpdecoy Otro" not in names, (
            "export must apply the same q filter as the page"
        )


# ---------------------------------------------------------------------------
# 2026-08-24 — el circuito de leads: primer mensaje visible y boton de WhatsApp
# ---------------------------------------------------------------------------

_WA_GENERICO = "conversar sobre propiedades en Paraguay"


def _wa_msg(html: str) -> str:
    """Devuelve el texto DECODIFICADO del `?text=` del href de wa.me.

    El `| urlencode` de Jinja escapa espacios como %20 y acentos como %XX; un
    assert por substring sobre el HTML crudo pasa verde sin probar nada — le
    paso a la version anterior de este test, que buscaba el generico sin
    escapar y por eso su `not in` era decorativo.
    """
    import re
    from urllib.parse import unquote_plus

    m = re.search(r"wa\.me/[^?]*\?text=([^'\"]*)", html)
    assert m is not None, "no hay href de wa.me con ?text= en la fila"
    return unquote_plus(m.group(1))


class TestBotonWhatsAppNombraLaPropiedad:
    """El guard de `lead_item.html` exigia `lead.property_id`, que para los
    leads de InfoCasas es NULL a proposito desde el fix del 71 % de matches
    Remax equivocados. La condicion nunca se cumplia y el `or lead.ic_title`
    no se llegaba a evaluar: el asesor abria WhatsApp con el texto generico
    aunque la fila estuviera mostrando el titulo de la propiedad.
    """

    @pytest.fixture
    async def ic_lead_sin_property_id(self, db):
        ref = "WAGRD1"
        title = "Duplex WA Probe en San Vicente"
        await db.execute(
            sa_text("DELETE FROM infocasas_properties WHERE infocasas_ref = :ref"),
            {"ref": ref},
        )
        await db.execute(
            sa_text(
                "INSERT INTO infocasas_properties "
                "(infocasas_id, infocasas_ref, title, city, is_active) "
                "VALUES ('WAGRD001', :ref, :title, 'San Vicente', true)"
            ),
            {"ref": ref, "title": title},
        )
        contact = _make_contact(
            db, f"+5959817{_suffix()}", source="infocasas",
            infocasas_ref=ref, property_id=None,
        )
        await db.commit()
        await db.refresh(contact)
        yield {"contact": contact, "title": title}
        await db.execute(
            sa_text("DELETE FROM infocasas_properties WHERE infocasas_ref = :ref"),
            {"ref": ref},
        )
        await db.commit()

    async def test_el_mensaje_precargado_nombra_la_propiedad(
        self, admin_client, ic_lead_sin_property_id, db
    ):
        """property_id NULL + ic_title presente -> el wa.me lleva el titulo."""
        contact = ic_lead_sin_property_id["contact"]
        assert contact.property_id is None, (
            "el fixture perdio su premisa: property_id tiene que ser NULL"
        )
        html = await _row_html(admin_client, contact)

        assert "wa.me/" in html, "el boton de WhatsApp no se renderizo"
        # El texto viaja urlencodeado adentro del href: se compara decodificado
        # para no acoplar el test al esquema de escape de Jinja.
        mensaje = _wa_msg(html)
        assert "Duplex WA Probe" in mensaje, (
            f"el mensaje precargado no nombra la propiedad — el guard sigue "
            f"mirando property_id. Mensaje: {mensaje!r}"
        )
        assert _WA_GENERICO not in mensaje, (
            "cayo al texto generico teniendo el titulo a mano"
        )

    async def test_sin_titulo_sigue_yendo_el_texto_generico(
        self, admin_client, db
    ):
        """La contracara: sin ningun titulo, el generico es lo correcto.

        Sin este caso, un guard que dijera `True` siempre pasaria el test de
        arriba igual.
        """
        contact = _make_contact(
            db, f"+5959817{_suffix()}", source="whatsapp", property_id=None,
        )
        await db.commit()
        await db.refresh(contact)
        html = await _row_html(admin_client, contact)
        assert _WA_GENERICO in _wa_msg(html), (
            "sin propiedad conocida el mensaje tiene que ser el generico"
        )


class TestPrimerMensajeVisible:
    """`first_message` se guardaba al 100 % y no lo leia nadie: ni el listado,
    ni la ficha, ni el xlsx. 366 de 865 traen HTML crudo de un tercero.
    """

    _HTML_SUCIO = (
        "\u00a1Has recibido una consulta reenviada!<br /><br />"
        "\u00bfEsta disponible a partir del 13 de septiembre?<br />"
        "\u00bfTienen generador?<br /><br />"
        "<img src=x onerror=alert(1)>"
    )

    async def test_la_ficha_del_contacto_muestra_la_consulta(
        self, admin_client, db
    ):
        contact = _make_contact(
            db, f"+5959817{_suffix()}", source="infocasas",
            first_message=self._HTML_SUCIO,
        )
        await db.commit()
        await db.refresh(contact)

        resp = await admin_client.get(f"/contacts/{contact.id}")
        assert resp.status_code == 200
        html = resp.text
        assert "Consulta del cliente" in html
        assert "13 de septiembre" in html, (
            "la pregunta concreta del cliente sigue invisible en la ficha"
        )
        assert "Tienen generador" in html

    async def test_el_html_de_un_tercero_no_llega_vivo_a_la_ficha(
        self, admin_client, db
    ):
        """XSS: el mensaje lo escribe un tercero (InfoCasas reenvia el bloque
        del cliente). Ni el <img> ni el onerror pueden salir como marcado."""
        contact = _make_contact(
            db, f"+5959817{_suffix()}", source="infocasas",
            first_message=self._HTML_SUCIO,
        )
        await db.commit()
        await db.refresh(contact)

        html = (await admin_client.get(f"/contacts/{contact.id}")).text
        assert "<img src=x" not in html, "el tag del tercero se renderizo vivo"
        assert "onerror=alert(1)" not in html, (
            "el handler llego sin escapar: es inyeccion"
        )
        bloque = html.split("Consulta del cliente")[1][:2000]
        assert "<br />" not in bloque, (
            "el <br> crudo se imprimio dentro del bloque de la consulta"
        )

    async def test_el_xlsx_lleva_la_consulta_en_texto_plano(
        self, admin_client, db
    ):
        from io import BytesIO

        from openpyxl import load_workbook

        marker = f"Wqmsg{_suffix()}"
        contact = _make_contact(
            db, f"+5959818{_suffix()}", name=f"{marker} Consulta",
            source="infocasas", status="interested",
            first_message=self._HTML_SUCIO,
        )
        await db.commit()
        await db.refresh(contact)

        resp = await admin_client.get(
            "/leads/export", params={"tab": "interesados", "q": marker},
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content), read_only=False)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert "Consulta del cliente" in headers, (
            "el xlsx sigue sin la columna del mensaje del cliente"
        )
        col = headers.index("Consulta del cliente") + 1
        celdas = [ws.cell(row=r, column=col).value or ""
                  for r in range(2, ws.max_row + 1)]
        texto = next((c for c in celdas if "septiembre" in c), None)
        assert texto is not None, "la consulta no llego al xlsx"
        assert "<br" not in texto and "<img" not in texto, (
            "el HTML crudo se exporto tal cual"
        )
