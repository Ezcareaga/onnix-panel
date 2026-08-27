"""Generate an .xlsx workbook with leads for the panel export button.

Pure presentation layer: takes already-fetched lead dicts (lead_repo shape)
and the last-messages map (contact_id -> [msg, ...]) and returns the
serialized workbook bytes. No DB access here — keeps the service testable
without a database fixture.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.tz import PYT, clean_description
from app.utils.money import miles, precio

EXPORT_COLUMNS = [
    "Nombre",
    "Teléfono",
    "Fuente",
    "Propiedad consultada",
    "URL propiedad",
    "Estado",
    "Primer contacto",
    "Último contacto",
    "Consulta del cliente",
]

_COLUMN_WIDTHS = [28, 18, 22, 50, 60, 16, 18, 18, 70]

# Excel corta la celda en 32.767 caracteres. Las consultas reales miden ~700,
# pero el recorte va acá y no en la confianza: una celda que revienta rompe el
# archivo entero, no una fila.
_MENSAJE_MAX = 4000


def _fmt_dt(dt: datetime | None) -> str:
    """Format a datetime in Paraguay time. Empty string when missing."""
    if dt is None:
        return ""
    if dt.tzinfo is not None:
        dt = dt.astimezone(PYT)
    return dt.strftime("%d/%m/%Y %H:%M")


def _source_label(lead: dict) -> str:
    """Map raw source to the exported label.

    InfoCasas leads split into 'infocasas_directo' / 'infocasas_reenviado'
    using the is_direct_ic flag the repo already computes.
    """
    source = lead.get("source")
    if source == "infocasas":
        return "infocasas_directo" if lead.get("is_direct_ic") else "infocasas_reenviado"
    return source or ""


def _property_summary(lead: dict) -> str:
    """One-cell summary: title + price + zone, from either properties or IC."""
    parts: list[str] = []
    title = lead.get("property_title") or lead.get("ic_title")
    if title:
        parts.append(title)

    if lead.get("property_price"):
        parts.append(precio(price_usd=lead["property_price"]))
    elif lead.get("ic_price_sale"):
        currency = "Gs" if lead.get("ic_currency_sale") == "PYG" else "USD"
        parts.append(f"{currency} {miles(lead['ic_price_sale'])} (Venta)")
    elif lead.get("ic_price_rent"):
        currency = "Gs" if lead.get("ic_currency_rent") == "PYG" else "USD"
        parts.append(f"{currency} {miles(lead['ic_price_rent'])} (Alquiler)")

    zona_bits = [
        lead.get("property_neighborhood") or lead.get("ic_city"),
        lead.get("property_city"),
    ]
    zona = ", ".join(z for z in zona_bits if z)
    if zona:
        parts.append(zona)

    return " — ".join(parts)


def _mensaje_cliente(lead: dict) -> str:
    """La consulta que escribió el cliente, en texto plano.

    `first_message` se guardaba al 100 % y no lo leía nadie: ni el listado, ni
    la ficha, ni el xlsx. Adentro hay preguntas concretas ("¿está disponible a
    partir del 13 de septiembre?", "¿tienen generador?").

    366 de los 865 mensajes de InfoCasas traen `<br />` crudo: son las
    reenviadas, que vienen envueltas en un bloque de aviso HTML. Se limpian con
    `clean_description` —el mismo filtro que usan las descripciones de las
    fichas— que reemplaza los `<br>` por saltos, **borra el resto de los tags**
    y decodifica entidades. Devuelve `str`, no `Markup`: openpyxl escribe texto
    y Jinja lo escaparía, así que no hay camino a inyección por ninguno de los
    dos lados.
    """
    limpio = clean_description(lead.get("first_message"))
    if len(limpio) > _MENSAJE_MAX:
        return limpio[:_MENSAJE_MAX] + "…"
    return limpio


def _property_url(lead: dict) -> str:
    """Return the property URL — prefers internal properties.url, falls back to ic_url."""
    return lead.get("property_url") or lead.get("ic_url") or ""


def build_leads_xlsx(leads: list[dict]) -> bytes:
    """Build the xlsx file as bytes. Pure function — no I/O."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_align = Alignment(horizontal="left", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    ws.append(EXPORT_COLUMNS)
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = _COLUMN_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for lead in leads:
        row = [
            lead.get("name") or "",
            lead.get("phone") or "",
            _source_label(lead),
            _property_summary(lead),
            _property_url(lead),
            lead.get("status") or "",
            _fmt_dt(lead.get("created_at")),
            _fmt_dt(lead.get("last_activity_at")),
            _mensaje_cliente(lead),
        ]
        ws.append(row)
        # Wrap the long property summary cell for readability
        last_row = ws.max_row
        ws.cell(row=last_row, column=4).alignment = wrap_align
        ws.cell(row=last_row, column=9).alignment = wrap_align
        # URL column: render as a clickable hyperlink when present
        url_cell = ws.cell(row=last_row, column=5)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="1F6FEB", underline="single")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(tab: str, source: str | None, status: str | None) -> str:
    """Build a content-disposition filename that reflects the active filter."""
    parts = ["leads", tab or "all"]
    if source:
        parts.append(source)
    if status:
        parts.append(status)
    parts.append(datetime.now(PYT).strftime("%Y%m%d_%H%M"))
    return "_".join(parts) + ".xlsx"
